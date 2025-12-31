# =========================================================
# 1️⃣ REQUIRED LIBRARIES (Importing the necessary tools for the program to run)
# =========================================================
# Selenium: Allows us to control the Chrome browser on the computer with our commands
from selenium import webdriver
# The "service" class needed to start ChromeDriver (the agent that opens the browser)
from selenium.webdriver.chrome.service import Service
# Methods used to locate buttons, text, input fields, etc. on the web page
# Examples: Find by ID, find by XPATH, find by CSS...
from selenium.webdriver.common.by import By
# To simulate keyboard presses (e.g., ENTER, TAB keys)
from selenium.webdriver.common.keys import Keys
# To wait for elements to appear on the page (e.g., continue when the search box is visible)
from selenium.webdriver.support.ui import WebDriverWait
# Expected conditions for waiting: "wait until button is clickable", "wait until text appears", etc.
from selenium.webdriver.support import expected_conditions as EC
# Automatically downloads and keeps ChromeDriver up to date
# No need to download it manually
from webdriver_manager.chrome import ChromeDriverManager
# To pause the program (for waiting periods)
import time
# To generate random numbers (we'll make wait times random to appear human-like)
import random
# Popular tool for handling data in tables and saving to Excel
import pandas as pd
# To open Excel files and modify font, column width, etc.
from openpyxl import load_workbook
# For formatting like bold text, alignment, etc.
from openpyxl.styles import Alignment, Font

# =========================================================
# 2️⃣ CONSTANT SETTINGS (Configurable settings are here)
# =========================================================
QUERY = "EEG Machine Learning" # Search term for Google Scholar
EXCEL_FILE = "eeg_2025_results.xlsx" # Filename where results will be saved
BASE_URL = "https://scholar.google.com/?hl=en" # Google Scholar English homepage
MAX_PAGES = 3 # How many pages to scan? (approximately 10 articles per page)
# =========================================================
# 3️⃣ HELPER FUNCTIONS (Small utility functions)
# =========================================================
def human_sleep(a=1.5, b=3):
    # Pauses for a random duration to mimic human behavior
    # Prevents Google from detecting us as a bot by avoiding fixed delays
    time.sleep(random.uniform(a, b)) # Sleeps for a random time between a and b seconds
def captcha_var_mi(driver):
    # Checks if a CAPTCHA or "unusual traffic" page is present
    # Google often shows messages like "unusual traffic" or "not a robot"
    try:
        # Look for common English CAPTCHA/unusual traffic indicators
        driver.find_element(By.XPATH, "//*[contains(text(),'unusual traffic') or contains(text(),'not a robot')]")
        return True # Yes, CAPTCHA or block detected
    except:
        return False # No CAPTCHA
def wait_for_captcha():
    # Pauses the program if CAPTCHA appears and asks the user for help
    print("\n⚠️ CAPTCHA detected!") # Warning message
    print(" • Do not close the browser.") # Keep the browser open
    print(" • Solve the CAPTCHA manually on the screen.") # Solve it by hand
    print(" • Refresh the page if necessary (F5 key).") # Refresh suggestion
    input(" >>> Press Enter after solving...") # Continue when user presses Enter
def click_xpath(driver, wait, xpath, description):
    # Clicks an element located by the given XPath
    print(f"👉 {description}") # Print what we're doing
    # Wait until the element is clickable (max 30 seconds)
    elem = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    # Scroll the element into the center of the view
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    human_sleep() # Human-like pause (1-3 seconds)
    elem.click() # Click it
    print(" ✅ Clicked.") # Success message
def safe_text(parent, css):
    # Safely extracts text (authors, abstract, etc.) from an article block
    # Returns empty string if not found (no error)
    try:
        return parent.find_element(By.CSS_SELECTOR, css).text
    except:
        return "" # Return empty if not found
def safe_link(parent):
    # Safely gets the title and link of an article
    # Returns "No title" and empty link if not found
    try:
        a_tag = parent.find_element(By.CSS_SELECTOR, "h3.gs_rt a") # Title link
        return a_tag.text, a_tag.get_attribute("href") # Title and URL
    except:
        return "No title", "" # Return defaults if missing
def format_excel(path):
    # Beautifies the Excel file for better readability
    print("🧹 Formatting Excel...")
    wb = load_workbook(path) # Open the file
    ws = wb.active # Get the active sheet
    # Make the header row (row 1) bold and centered
    for cell in ws[1]:
        cell.font = Font(bold=True) # Bold font
        cell.alignment = Alignment(horizontal="center") # Center align
    # Manually set column widths
    widths = [6, 55, 40, 95, 60] # Order, Title, Authors, Abstract, Link
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width # A=1, B=2, etc.
    # Enable text wrapping for the Abstract column (column D)
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top") # Wrap text
    wb.save(path) # Save changes
    print(" ✨ Excel ready.")
# =========================================================
# 4️⃣ START THE BROWSER (Launching Chrome)
# =========================================================
print("🚀 Starting browser...")
# Configure Chrome options (to hide bot detection)
options = webdriver.ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled") # Hide automation traces
options.add_experimental_option("excludeSwitches", ["enable-automation"]) # Additional hiding
# Recommended settings for extra stealth and human-like behavior
options.add_argument("--no-sandbox") # Needed in Linux/server environments + some stealth
options.add_argument("--disable-dev-shm-usage") # Prevents memory issues, reduces bot traces
options.add_argument("--disable-gpu") # Disable GPU (useful in headless mode)
options.add_argument("--disable-infobars") # Hide info bars in older versions (still sometimes useful)
#options.add_argument("--start-maximized") # Start browser maximized (more natural)
# Change User-Agent (very effective on sites like Google Scholar)
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")
# Launch Chrome (ChromeDriver will be downloaded and used automatically)
driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()), # Auto-install ChromeDriver
    options=options
)
# Smart wait object that can wait up to 30 seconds
wait = WebDriverWait(driver, 30)
# =========================================================
# 5️⃣ MAIN PROGRAM FLOW
# =========================================================
try: # Catch any errors
    # Open Google Scholar homepage
    print("📂 Opening Google Scholar...")
    driver.get(BASE_URL) # Navigate to the URL
    human_sleep(3, 5) # Wait for full page load
    # Check for CAPTCHA on initial load
    if captcha_var_mi(driver):
        wait_for_captcha()

    # Perform search with direct URL: 2025+, sorted by date (newest first)
    print(f"🔍 Searching for: '{QUERY}'")
    SEARCH_URL = f"https://scholar.google.com/scholar?hl=en&q={QUERY.replace(' ', '+')}&as_ylo=2025&scisbd=1"
    driver.get(SEARCH_URL)
    print("🕒 'Since 2025' filter and 'Sort by date' applied directly via URL")
    human_sleep(2, 4) # Wait for results

    # Verify results are loaded (try up to 10 times)
    for _ in range(10):
        if captcha_var_mi(driver):
            wait_for_captcha()
            break
        try:
            # Check if an article block exists
            driver.find_element(By.CSS_SELECTOR, "div.gs_r.gs_scl")
            print(" ✅ Results loaded.")
            break
        except:
            time.sleep(2) # Wait a bit more if not ready

    # =========================================================
    # PAGINATION: NAVIGATE PAGES (We'll scan the first 3 pages)
    # =========================================================
    all_data = [] # List to store all collected articles
    current_page = 1 # Current page number
    while current_page <= MAX_PAGES: # Loop through MAX_PAGES
        print(f"\n📄 Processing page {current_page}...")
        if captcha_var_mi(driver):
            wait_for_captcha()
        # Find all article blocks on the current page
        rows = driver.find_elements(By.CSS_SELECTOR, "div.gs_r.gs_scl")
        print(f" → {len(rows)} articles found.")
        # Process each article
        for row in rows:
            title, link = safe_link(row) # Get title and link
            authors = safe_text(row, "div.gs_a") # Get authors
            abstract = safe_text(row, "div.gs_rs") # Get abstract
            # Add article to the list
            all_data.append([
                len(all_data) + 1, # Row number (1,2,3...)
                title, # Title
                authors, # Authors
                abstract, # Abstract
                link # Link
            ])
        # Move to next page if needed
        if current_page < MAX_PAGES:
            clicked = False # Did we successfully click next?
            retry_count = 0 # How many attempts?
            while retry_count < 3 and not clicked:
                try:
                    # Find the "Next" button (English interface)
                    next_button = wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "button[aria-label='Next']"))
                    )
                    # Check if button is enabled
                    if "gs_btn_dis" not in next_button.get_attribute("class"):
                        # Scroll button into view
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", next_button)
                        human_sleep(1, 2) # Short pause
                        print(f"➡️ Clicking Next button (attempt {retry_count+1})...")
                        driver.execute_script("arguments[0].click();", next_button) # Click via JavaScript
                        human_sleep(7, 12) # Wait for new page to load
                        # Verify new page loaded
                        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.gs_r.gs_scl")))
                        clicked = True # Success
                        current_page += 1 # Increment page
                    else:
                        print(" Button disabled, retrying...")
                        human_sleep(2, 4)
                        retry_count += 1
                except Exception as e:
                    print(f" Button not found (attempt {retry_count+1}): {e}")
                    human_sleep(2, 4)
                    retry_count += 1
            if not clicked:
                print("⚠️ Next button could not be clicked after 3 attempts. Likely on the last page.")
                break # No more pages
        else:
            print("✅ Reached maximum page limit.")
            break
    # =========================================================
    # SAVE DATA TO EXCEL
    # =========================================================
    # Convert collected data to DataFrame
    df = pd.DataFrame(
        all_data,
        columns=["Order", "Title", "Authors", "Abstract", "Link"]
    )
    # Save to Excel (without index column)
    df.to_excel(EXCEL_FILE, index=False)
    # Format the Excel file nicely
    format_excel(EXCEL_FILE)
    print(f"\n🎯 TOTAL {len(all_data)} articles collected → {EXCEL_FILE}")
    print(f"\nALL OPERATIONS COMPLETED.")
except Exception as e:
    # Print any critical error
    print(f"\n❌ Critical error: {e}")
    if captcha_var_mi(driver):
        wait_for_captcha() # Alert user if CAPTCHA appears
finally:
    # Close browser when done
    print("\n🌙 Browser will close in 10 seconds...")
    time.sleep(10) # Pause to view results
    driver.quit() # Close browser
    print("👋 Browser closed.")